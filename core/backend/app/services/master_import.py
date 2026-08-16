"""Import an operator's fleet spreadsheet into the ``trucks`` master table.

Default source: ``sources/*.xlsx`` -- the "DAFTAR KENDARAAN / UNIT" sheet
PT. CK - BIB maintains (header on row 9, data from row 12, hull always
``"HD" + 4 digits`` in ``No. Lambung Kend. / Unit``). Other contractors'
sheets pass their own path explicitly; each contractor's fleet is added
alongside what's already imported rather than replacing it (see
``import_master``'s ``replace`` flag). The header row is located by content
rather than hardcoded, so extra title rows or a contractor's own column
wording ("Tahun" vs "Tahun Kend. / Unit", "Keterangan" vs "Status") still
import; only ``COLUMN_HINTS`` needs to know the label.

    uv run python -m app.services.master_import                 # default sources/ file
    uv run python -m app.services.master_import path/to.xlsx    # explicit
    uv run python -m app.services.master_import --dry-run       # parse and report only
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from app.core.config import ROOT
from app.repositories import truck_master_repo

SOURCES_DIR = ROOT / "sources"

# Header label -> our field. Matched case-insensitively on a prefix, because the
# sheet's labels carry trailing units and slashes that vary between revisions
# -- and, across contractors, entirely different wording ("Tahun" vs "Tahun
# Kend. / Unit", "Keterangan" vs "Status"). Where a sheet has both a generic and
# a specific column for the same field (CK-BIB has both "Keterangan" and a
# separate "Status"), the later column in reading order wins -- see
# `parse_workbook`'s record assembly.
COLUMN_HINTS = {
    "perusahaan": "contractor",
    "jenis kendaraan": "unit_type",
    "merek": "brand",
    "type": "model_type",
    "no. lambung": "hull_id",
    "tahun": "year",
    "keterangan": "status",
    "status": "status",
}

# The hull's digits, whatever their length -- CK-BIB always paints exactly 4,
# but other contractors' sheets (e.g. PPA) mix 2-5 digit unit numbers. A hull_id
# with anything other than exactly one digit run is unparseable, not guessed.
_HULL_DIGITS = re.compile(r"(\d+)")


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _find_header_row(ws, max_scan: int = 30) -> tuple[int, dict[int, str]]:
    """Locate the header row and map column index -> field name."""
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        labels = {i: str(c).strip().lower() for i, c in enumerate(row) if c}
        if not any("lambung" in v for v in labels.values()):
            continue
        mapping: dict[int, str] = {}
        for idx, label in labels.items():
            for hint, field in COLUMN_HINTS.items():
                if label.startswith(hint):
                    mapping[idx] = field
                    break
        if "hull_id" in mapping.values():
            return row_idx, mapping
    raise ValueError(
        "could not find a header row containing 'No. Lambung' in the first "
        f"{max_scan} rows -- is this the right sheet?"
    )


def parse_workbook(path: Path, sheet: str | None = None) -> tuple[list[dict], list[str]]:
    """Parse the spreadsheet into master rows. Returns ``(rows, warnings)``."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    header_row, mapping = _find_header_row(ws)
    rows: list[dict] = []
    warnings: list[str] = []
    seen_codes: dict[str, str] = {}

    for excel_row, values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        record = {field: _clean(values[idx]) for idx, field in mapping.items()
                  if idx < len(values)}
        hull_id = record.get("hull_id")
        if not hull_id or hull_id == "-":
            continue  # spacer / subtotal row

        digits = _HULL_DIGITS.findall(hull_id)
        if len(digits) != 1:
            warnings.append(
                f"row {excel_row}: hull {hull_id!r} has "
                f"{'no digits' if not digits else 'more than one digit run'} -- skipped"
            )
            continue
        hull_code = digits[0]

        if hull_code in seen_codes:
            warnings.append(
                f"row {excel_row}: hull code {hull_code} ({hull_id}) collides with "
                f"{seen_codes[hull_code]} -- skipped, the code must be unique to match OCR"
            )
            continue
        seen_codes[hull_code] = hull_id

        year = record.get("year")
        try:
            year_val = int(float(year)) if year not in (None, "-") else None
        except (TypeError, ValueError):
            year_val = None

        rows.append({
            "hull_id": hull_id,
            "hull_code": hull_code,
            "contractor": record.get("contractor"),
            "unit_type": record.get("unit_type"),
            "brand": record.get("brand"),
            "model_type": record.get("model_type"),
            "year": year_val,
            "status": record.get("status"),
        })

    return rows, warnings


def default_source() -> Path | None:
    """The newest ``.xlsx`` in ``sources/``, if any."""
    if not SOURCES_DIR.is_dir():
        return None
    sheets = sorted(
        (p for p in SOURCES_DIR.glob("*.xlsx") if p.is_file()),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return sheets[0] if sheets else None


def import_master(path: Path | None = None, *, replace: bool = False) -> dict:
    """Parse and persist the master registry. Returns a summary.

    Without ``replace``, this adds to whatever is already in ``trucks`` --
    multiple contractors' fleets coexist in one table. Because ``hull_code``
    must stay globally unique to match OCR, a parsed row whose code already
    belongs to a *different* hull_id is skipped (not overwritten, not allowed
    to abort the whole import) with a warning; the same code re-imported for
    the same hull_id is treated as an update, same as within one sheet.
    """
    source = path or default_source()
    if source is None or not source.is_file():
        raise SystemExit(f"no spreadsheet found (looked in {SOURCES_DIR})")

    rows, warnings = parse_workbook(source)
    if not rows:
        raise SystemExit(f"{source.name}: no usable unit rows found")

    if replace:
        truck_master_repo.clear()
    else:
        existing = {t["hull_code"]: t["hull_id"] for t in truck_master_repo.list_all()}
        kept = []
        for r in rows:
            clash = existing.get(r["hull_code"])
            if clash is not None and clash != r["hull_id"]:
                warnings.append(
                    f"hull code {r['hull_code']} ({r['hull_id']}) is already registered "
                    f"as {clash} -- skipped, the code must be unique to match OCR"
                )
                continue
            kept.append(r)
        rows = kept

    result = truck_master_repo.upsert_many(rows)
    return {"source": source.name, "parsed": len(rows), "warnings": warnings, **result}


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    flags = {a for a in sys.argv[1:] if a.startswith("-")}
    source = Path(args[0]) if args else default_source()

    if "--dry-run" in flags:
        if source is None:
            raise SystemExit(f"no spreadsheet found in {SOURCES_DIR}")
        rows, warnings = parse_workbook(source)
        print(f"{source.name}: parsed {len(rows)} units (dry run, nothing written)")
        for w in warnings:
            print(f"  WARN {w}")
        for r in rows[:5]:
            print(f"  {r['hull_id']:<10} code={r['hull_code']} {r['unit_type']} "
                  f"{r['brand']} {r['model_type']} {r['year']} {r['status']}")
        return

    summary = import_master(source, replace="--replace" in flags)
    print(
        f"{summary['source']}: parsed {summary['parsed']} units -> "
        f"{summary['inserted']} inserted, {summary['updated']} updated "
        f"({truck_master_repo.count()} in master)"
    )
    for w in summary["warnings"]:
        print(f"  WARN {w}")


if __name__ == "__main__":
    main()
