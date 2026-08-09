"""The truck master registry: import, persistence, and edge-ingestion wiring."""

from __future__ import annotations

import pytest

from app.repositories import truck_master_repo as tm
from app.services import master_import

PYTEST_HULL = "ZZ 9911"
PYTEST_CODE = "9911"


def _row(hull_id=PYTEST_HULL, hull_code=PYTEST_CODE, **over):
    row = {
        "hull_id": hull_id, "hull_code": hull_code, "contractor": "PYTEST CO",
        "unit_type": "OHT", "brand": "CATERPILLAR", "model_type": "777",
        "year": 2020, "status": "Layak",
    }
    row.update(over)
    return row


@pytest.fixture(autouse=True)
def _cleanup():
    yield
    from app.core.database import connect
    conn = connect()
    try:
        conn.execute("DELETE FROM trucks WHERE contractor = 'PYTEST CO'")
        conn.commit()
    finally:
        conn.close()


# --- repository --------------------------------------------------------------

def test_upsert_inserts_then_updates():
    before = tm.count()
    assert tm.upsert_many([_row()]) == {"inserted": 1, "updated": 0}
    assert tm.count() == before + 1

    # Same hull_id again is an update, not a duplicate -- a corrected sheet can
    # simply be re-imported.
    assert tm.upsert_many([_row(model_type="773E")]) == {"inserted": 0, "updated": 1}
    assert tm.count() == before + 1
    assert tm.get_by_hull_id(PYTEST_HULL)["model_type"] == "773E"


def test_lookup_by_code_and_id():
    tm.upsert_many([_row()])
    assert tm.get_by_hull_code(PYTEST_CODE)["hull_id"] == PYTEST_HULL
    assert tm.get_by_hull_id(PYTEST_HULL)["hull_code"] == PYTEST_CODE
    assert tm.get_by_hull_code("0000") is None


def test_hull_code_appears_in_candidate_set():
    tm.upsert_many([_row()])
    assert PYTEST_CODE in tm.all_hull_codes()


# --- spreadsheet parsing -----------------------------------------------------

def test_parses_the_real_operator_sheet():
    """The shipped sheet must import cleanly -- it is the production source."""
    source = master_import.default_source()
    if source is None:
        pytest.skip("no spreadsheet in sources/")

    rows, warnings = master_import.parse_workbook(source)
    assert rows, "no unit rows parsed"
    assert not warnings, f"unexpected parse warnings: {warnings[:3]}"

    # Every hull yields exactly one 4-digit code, and those codes are unique --
    # the property the whole matching strategy depends on.
    codes = [r["hull_code"] for r in rows]
    assert all(len(c) == 4 and c.isdigit() for c in codes)
    assert len(set(codes)) == len(codes), "duplicate hull codes would break matching"

    for r in rows:
        assert r["hull_id"] and r["hull_code"] in r["hull_id"].replace(" ", "")


def test_header_row_is_found_by_content_not_position():
    """Extra title rows above the header must not break the import."""
    source = master_import.default_source()
    if source is None:
        pytest.skip("no spreadsheet in sources/")
    import openpyxl

    ws = openpyxl.load_workbook(source, data_only=True)[
        openpyxl.load_workbook(source, data_only=True).sheetnames[0]
    ]
    header_row, mapping = master_import._find_header_row(ws)
    assert header_row > 1, "the sheet has title rows above the header"
    assert "hull_id" in mapping.values()


# --- edge ingestion resolves through the matcher -----------------------------

def test_edge_crossing_resolves_a_noisy_read_to_the_master(client, edge_camera, auth_headers):
    """A device reporting '9911' is stored as the master's 'ZZ 9911'."""
    import json
    import uuid

    tm.upsert_many([_row()])
    key = str(uuid.uuid4())
    payload = {
        "camera_code": edge_camera["camera_code"],
        "detected_at": "2026-08-02T14:31:02Z", "window_sec": 5.8,
        "hull_id": PYTEST_CODE, "confidence": 0.94, "read_count": 3,
        "votes": [{"text": PYTEST_CODE, "count": 3, "avg_ocr_conf": 0.9}],
    }
    r = client.post(
        "/api/edge/crossings",
        headers={**auth_headers, "Idempotency-Key": key},
        data={"payload": json.dumps(payload)},
        files={"snapshot": ("c.jpg", b"\xff\xd8j\xff\xd9", "image/jpeg")},
    )
    assert r.status_code == 201

    from app.core.database import connect
    conn = connect()
    try:
        stored = conn.execute(
            "SELECT voted_hull_id FROM video_results WHERE idempotency_key = ?", (key,)
        ).fetchone()[0]
    finally:
        conn.close()
    assert stored == PYTEST_HULL, "the reading was not resolved to the master hull id"


def test_edge_crossing_with_unregistered_read_is_unknown(client, edge_camera, auth_headers):
    """An unmatched read is stored as UNKNOWN, never as an invented hull."""
    import json
    import uuid

    from app.core.config import UNIDENTIFIED_HULLS

    key = str(uuid.uuid4())
    payload = {
        "camera_code": edge_camera["camera_code"],
        "detected_at": "2026-08-02T14:31:02Z", "window_sec": 5.8,
        "hull_id": "7777", "confidence": 0.94, "read_count": 3,
        "votes": [{"text": "7777", "count": 3, "avg_ocr_conf": 0.9}],
    }
    r = client.post(
        "/api/edge/crossings",
        headers={**auth_headers, "Idempotency-Key": key},
        data={"payload": json.dumps(payload)},
        files={"snapshot": ("c.jpg", b"\xff\xd8j\xff\xd9", "image/jpeg")},
    )
    assert r.status_code == 201

    from app.core.database import connect
    conn = connect()
    try:
        stored = conn.execute(
            "SELECT voted_hull_id FROM video_results WHERE idempotency_key = ?", (key,)
        ).fetchone()[0]
    finally:
        conn.close()
    assert stored in UNIDENTIFIED_HULLS
