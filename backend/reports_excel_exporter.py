import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.reports_logic import calculate_shift_summary
from backend import database

def generate_reconciliation_excel() -> io.BytesIO:
    wb = Workbook()
    
    # 1. Daily Statistics & Shift Summary
    ws1 = wb.active
    ws1.title = "Daily & Shift Stats"
    
    summary = calculate_shift_summary()
    
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1E3A8A")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Daily Statistics Section
    ws1.cell(row=1, column=1, value="Daily Checkpoint Volume").font = title_font
    ws1.row_dimensions[1].height = 25
    
    ws1.cell(row=3, column=1, value="Date").font = header_font
    ws1.cell(row=3, column=1).fill = header_fill
    ws1.cell(row=3, column=2, value="Crossings Volume").font = header_font
    ws1.cell(row=3, column=2).fill = header_fill
    ws1.row_dimensions[3].height = 20
    
    r = 4
    for date_str, vol in summary.get("date_distribution", {}).items():
        ws1.cell(row=r, column=1, value=date_str).font = regular_font
        ws1.cell(row=r, column=2, value=vol).font = regular_font
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border
        r += 1
        
    # Shift Distribution Section
    r += 2
    ws1.cell(row=r, column=1, value="Shift Time Slot Distribution").font = title_font
    r += 2
    
    ws1.cell(row=r, column=1, value="Shift Slot").font = header_font
    ws1.cell(row=r, column=1).fill = header_fill
    ws1.cell(row=r, column=2, value="Crossings Count").font = header_font
    ws1.cell(row=r, column=2).fill = header_fill
    ws1.row_dimensions[r].height = 20
    
    r += 1
    for slot, cnt in summary.get("shift_distribution", {}).items():
        ws1.cell(row=r, column=1, value=slot).font = regular_font
        ws1.cell(row=r, column=2, value=cnt).font = regular_font
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border
        r += 1
        
    # Truck Ritase Section
    r += 2
    ws1.cell(row=r, column=1, value="OHT Ritase & Passages Summary").font = title_font
    r += 2
    
    ws1.cell(row=r, column=1, value="Truck Hull ID").font = header_font
    ws1.cell(row=r, column=1).fill = header_fill
    ws1.cell(row=r, column=2, value="Crossings").font = header_font
    ws1.cell(row=r, column=2).fill = header_fill
    ws1.cell(row=r, column=3, value="Completed Ritase (Trips)").font = header_font
    ws1.cell(row=r, column=3).fill = header_fill
    ws1.row_dimensions[r].height = 20
    
    r += 1
    for hid, crossings in summary.get("crossings_per_truck", {}).items():
        ritase = summary.get("completed_ritase", {}).get(hid, 0)
        ws1.cell(row=r, column=1, value=hid).font = regular_font
        ws1.cell(row=r, column=2, value=crossings).font = regular_font
        ws1.cell(row=r, column=3, value=ritase).font = regular_font
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border
        ws1.cell(row=r, column=3).border = thin_border
        r += 1
        
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # 2. Subcontractor Compliance Summaries
    ws2 = wb.create_sheet(title="Subcontractor Compliance")
    ws2.cell(row=1, column=1, value="Subcontractor Shift Compliance Summary").font = title_font
    ws2.row_dimensions[1].height = 25
    
    headers2 = [
        "Subcontractor", "Actual Ritase/Hr", "Target Ritase/Hr", 
        "Compliance %", "Active Fleet Size", "Min Fleet Size", "Fleet Compliance %"
    ]
    for c_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws2.row_dimensions[3].height = 20
    
    r = 4
    for contractor, data in summary.get("compliance", {}).items():
        ws2.cell(row=r, column=1, value=contractor).font = bold_font
        ws2.cell(row=r, column=2, value=round(data.get("actual_ritase_hr", 0.0), 2)).font = regular_font
        ws2.cell(row=r, column=3, value=round(data.get("target_ritase_hr", 0.0), 2)).font = regular_font
        ws2.cell(row=r, column=4, value=f"{round(data.get('compliance_pct', 0.0), 1)}%").font = bold_font
        ws2.cell(row=r, column=5, value=data.get("active_fleet_size", 0)).font = regular_font
        ws2.cell(row=r, column=6, value=data.get("min_fleet_size", 0)).font = regular_font
        ws2.cell(row=r, column=7, value=f"{round(data.get('fleet_compliance_pct', 0.0), 1)}%").font = bold_font
        
        for col_idx in range(1, 8):
            ws2.cell(row=r, column=col_idx).border = thin_border
        r += 1
        
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 16)

    # 3. Discrepancy Logs
    ws3 = wb.create_sheet(title="Discrepancy Logs")
    ws3.cell(row=1, column=1, value="System Discrepancies & Alerts Log").font = title_font
    ws3.row_dimensions[1].height = 25
    
    headers3 = ["Timestamp", "Device / Hull ID", "Location / Lane", "Alert Type", "Severity", "Details"]
    for c_idx, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=3, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws3.row_dimensions[3].height = 20
    
    r = 4
    for d in summary.get("discrepancies", []):
        ws3.cell(row=r, column=1, value=d.get("timestamp", "")).font = regular_font
        ws3.cell(row=r, column=2, value=d.get("hull_id", "")).font = bold_font
        ws3.cell(row=r, column=3, value=d.get("lane", "")).font = regular_font
        ws3.cell(row=r, column=4, value=d.get("type", "")).font = bold_font
        ws3.cell(row=r, column=5, value=str(d.get("severity", "")).upper()).font = bold_font
        ws3.cell(row=r, column=6, value=d.get("details", "")).font = regular_font
        
        sev = str(d.get("severity", "")).lower()
        if sev == "high":
            ws3.cell(row=r, column=5).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        elif sev == "medium":
            ws3.cell(row=r, column=5).fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            
        for col_idx in range(1, 7):
            ws3.cell(row=r, column=col_idx).border = thin_border
        r += 1
        
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 16)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
